# day3_dashboard.py
import os
import json
import logging
from datetime import datetime
from flask import Flask, jsonify, render_template, request
import gspread
from google.oauth2.service_account import Credentials

# Optional: read local xlsx
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

# ---------- Logging ----------
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("autopulse")

# ---------- Google scopes ----------
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets.readonly",
    "https://www.googleapis.com/auth/drive.readonly",
    # For writing from master to sheet we need spreadsheets scope
    "https://www.googleapis.com/auth/spreadsheets"
]

# ---------- Config from env ----------
SHEET_ID = os.getenv("SHEET_ID", "").strip()
SHEET_NAME = os.getenv("SHEET_NAME", "Sheet1").strip()
MASTER_XLSX_PATH = os.getenv("MASTER_XLSX_PATH", "sales_data.xlsx").strip()
TEST_MODE = os.getenv("TEST_MODE", "0") == "1"

# ---------- Credential loader ----------
def load_credentials():
    """Load service account credentials either from GOOGLE_CREDS env or local file."""
    creds_json = os.getenv("GOOGLE_CREDS")
    if creds_json:
        try:
            # try to parse either raw JSON or escaped string
            try:
                info = json.loads(creds_json)
            except json.JSONDecodeError:
                fixed = creds_json.replace("'", '"').replace("\\n", "")
                info = json.loads(fixed)
            return Credentials.from_service_account_info(info, scopes=SCOPES)
        except Exception:
            logger.exception("Failed to parse GOOGLE_CREDS")
            raise

    # fallback to service_account.json next to this file
    local = os.path.join(os.path.dirname(__file__), "service_account.json")
    if os.path.exists(local):
        return Credentials.from_service_account_file(local, scopes=SCOPES)

    raise ValueError("Missing Google credentials. Set GOOGLE_CREDS or add service_account.json")

def get_gspread_client():
    creds = load_credentials()
    return gspread.authorize(creds)

# ---------- Helpers ----------
def parse_number(value):
    """Safe numeric parser that strips commas, currency symbols, blanks."""
    try:
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        s = str(value).strip()
        if s == "":
            return 0.0
        s = s.replace(",", "").replace("₹", "").replace("Rs", "").strip()
        return float(s)
    except Exception:
        return 0.0

def detect_last_numeric_columns(rows, header_row):
    """
    rows: list of value-lists for each data row (without header)
    header_row: list of header labels
    Returns (last_col_index, prev_col_index) where indexes are 0-based.
    Assumes first two columns are category and revenue (or similar) and numeric columns start at index>=2.
    """
    ncols = len(header_row)
    numeric_cols = []
    # examine columns from index 2 onwards (0-based)
    for c in range(2, ncols):
        count_numeric = 0
        for r in rows:
            if c < len(r):
                if parse_number(r[c]) != 0.0:
                    count_numeric += 1
        if count_numeric > 0:
            numeric_cols.append(c)
    if not numeric_cols:
        return (None, None)
    last = numeric_cols[-1]
    prev = numeric_cols[-2] if len(numeric_cols) > 1 else None
    return (last, prev)

# ---------- Fetch sheet data ----------
def fetch_sheet_data():
    """
    Return list of dicts:
    [{ category, latest, previous, growth, date }, ...]
    """
    if not SHEET_ID:
        logger.warning("SHEET_ID not set")
        return []

    try:
        gc = get_gspread_client()
        sh = gc.open_by_key(SHEET_ID)
        # Get worksheet by name (must match tab)
        ws = sh.worksheet(SHEET_NAME)

        all_values = ws.get_all_values()
        if not all_values or len(all_values) < 2:
            logger.info("Sheet empty or only header")
            return []

        header = all_values[0]
        rows = all_values[1:]

        last_col, prev_col = detect_last_numeric_columns(rows, header)
        if last_col is None:
            logger.info("No numeric columns detected")
            return []

        data = []
        for r in rows:
            category = r[0] if len(r) > 0 and r[0].strip() != "" else "Unknown"
            latest = parse_number(r[last_col]) if len(r) > last_col else 0.0
            previous = parse_number(r[prev_col]) if prev_col is not None and len(r) > prev_col else 0.0
            growth = None
            if previous and previous != 0:
                growth = round((latest - previous) / previous * 100, 2)
            data.append({
                "category": category,
                "latest": latest,
                "previous": previous,
                "growth": growth,
                "date": header[last_col]
            })
        return data
    except Exception:
        logger.exception("Error fetching Google Sheet data")
        return []

# ---------- Sync master -> Google Sheet (append next column) ----------
def sync_from_master():
    """
    Read MASTER_XLSX_PATH (sheet must have first column = category)
    and append the next values column to the Google Sheet (keeping existing).
    - Only writes the next column (header + values)
    - Does not remove old data
    Returns dict status.
    """
    if not SHEET_ID:
        return {"ok": False, "error": "SHEET_ID not configured"}

    if not OPENPYXL_AVAILABLE:
        return {"ok": False, "error": "openpyxl not available in runtime. Add to requirements."}

    if not os.path.exists(MASTER_XLSX_PATH):
        return {"ok": False, "error": f"Master file not found at {MASTER_XLSX_PATH}"}

    try:
        wb = load_workbook(MASTER_XLSX_PATH, data_only=True)
        ws_master = wb.active  # assume master data in first sheet
        # read entire master into list of rows
        master_rows = []
        for row in ws_master.iter_rows(values_only=True):
            master_rows.append(list(row))
        if not master_rows or len(master_rows) < 2:
            return {"ok": False, "error": "Master dataset seems empty or too small"}

        # master header and rows
        master_header = master_rows[0]
        master_data_rows = master_rows[1:]

        # Map category -> values for latest column in master
        # assume master columns are: category, baseline,..., next_day_value (final column)
        latest_col_index = len(master_header) - 1
        master_map = {}
        for r in master_data_rows:
            if len(r) == 0:
                continue
            cat = str(r[0]).strip() if r[0] is not None else "Unknown"
            val = r[latest_col_index] if latest_col_index < len(r) else 0
            master_map[cat] = parse_number(val)

        # connect and write to google sheet
        gc = get_gspread_client()
        sh = gc.open_by_key(SHEET_ID)
        ws = sh.worksheet(SHEET_NAME)

        # get header of sheet and find next free column index (1-based)
        sheet_header = ws.row_values(1)
        next_col_index = len(sheet_header) + 1  # append at end

        # write header value (timestamp or master header)
        header_value = master_header[latest_col_index] if master_header[latest_col_index] else datetime.now().isoformat()
        ws.update_cell(1, next_col_index, str(header_value))

        # iterate over rows in sheet and update the column with master_map values by category
        sheet_rows = ws.get_all_values()[1:]  # all data rows
        for i, srow in enumerate(sheet_rows, start=2):  # start=2 because sheet row 1 header
            cat = srow[0].strip() if len(srow) > 0 and srow[0] else "Unknown"
            v = master_map.get(cat, None)
            if v is None:
                # if master doesn't have this category, skip (do not overwrite)
                continue
            ws.update_cell(i, next_col_index, v)

        return {"ok": True, "written_column": next_col_index, "header": header_value}
    except Exception:
        logger.exception("sync_from_master failed")
        return {"ok": False, "error": "exception during sync"}

# ---------- Flask app ----------
app = Flask(__name__, static_folder="static", template_folder="templates")

@app.route("/")
def home():
    return render_template("index.html")  # keep your existing index.html

@app.route("/api/data")
def api_data():
    data = fetch_sheet_data()
    if not data:
        # fallback sample
        return jsonify({
            "data": [
                {"category":"Electronics","latest":120000,"previous":100000,"growth":20.0},
                {"category":"Books","latest":45000,"previous":48000,"growth":-6.25}
            ],
            "updated_at": datetime.utcnow().isoformat()
        })
    return jsonify({"data": data, "updated_at": datetime.utcnow().isoformat()})

# Optional endpoint to trigger a sync from master -> Google Sheets (POST for safety)
@app.route("/api/sync", methods=["POST"])
def api_sync():
    # If you want to limit this endpoint, check a secret header or env token here.
    result = sync_from_master()
    return jsonify(result)

if __name__ == "__main__":
    port = int(os.getenv("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
